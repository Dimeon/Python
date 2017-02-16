'''from openpyxl import load_workbook
wb = load_workbook('C:/Project/ReqParser/Implementation_Planing_v15.xlsx')
print(wb.get_sheet_names())
sheet = wb['SWRS-AUDGenIV-V24']
val = sheet['A9':'A40']
print(val.value)
wb.save('C:/Project/ReqParser/Implementation_Planing_v15.xlsx')'''

import openpyxl
wb = openpyxl.load_workbook('C:/Project/ReqParser/Implementation_Planing_v15.xlsx')
sheet = wb['SWRS-AUDGenIV-V24']
excel_reqs = []

#here you iterate over the rows in the specific column
for row in range(9,sheet.max_row+1):
    for column in "A":  #Here you can add or reduce the columns
        cell_name = "{}{}".format(column, row)
        excel_reqs.append(sheet[cell_name].value)

        #sheet[cell_name].value # the value of the specific cell
        print(cell_name)
        print(sheet[cell_name].value)
print(excel_reqs)



