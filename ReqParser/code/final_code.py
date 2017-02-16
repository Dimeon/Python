import openpyxl
import xml.etree.ElementTree as ET


wb = openpyxl.load_workbook('C:/Project/Python/ReqParser/Reports/Implementation_Planing_v19.xlsx')
sheet = wb['SWRS-AUDGenIV-V24']
excel_req_info = []  # [ (row_nr, req), (row_nr, req), .. ]

#here you iterate over the rows in the specific column
for row in range(9, sheet.max_row+1):
    for column in "A":  #Here you can add or reduce the columns
        cell_name = "{}{}".format(column, row)
        excel_req_info.append(
            (row, sheet[cell_name].value)
        )


tree = ET.ElementTree(file="C:/Project/Python/ReqParser/Temp/the_latest/Eng8_SW_Tests.xml")
root = tree.getroot()

idents = [el.text for el in root.iterfind('.//testcase/ident')]
descs = [el.text for el in root.iterfind('.//testcase/description')]

d = {id_: desc for id_, desc in zip(idents, descs)}

print(d)
from typing import List

def find_trace(trace: str) -> List[str]:
    l = []
    for k, v in d.items():
        if trace in v:
            l.append(k)
            print(l)
    return l

for row_nr, req in excel_req_info:
    trace_id = find_trace(req)
    trace_id = str(trace_id)[2:-2]
    print(row_nr, req, trace_id)
    sheet.cell(row=row_nr, column=13).value = trace_id

wb.save('C:/Project/Python/ReqParser/output/new_big_file.xlsx')
