import openpyxl
import xml.etree.ElementTree as ET
from pprint import pprint
import re
import json
import sys


def excel_file_parse():
    xlsx_file = 'C:/Project/ReqParser/Reports/Implementation_Planing_v19.xlsx'
    return xlsx_file


def sheet_in_excel(xlsx_sheet):
    xlsx_sheet = 'SWRS-AUDGenIV-V24'
    return xlsx_sheet


def xml_file_parse(xml_file):
    xml_file = "C:/Project/ReqParser/Reports/Eng8_SW_Tests_XLSX.xml"
    return xml_file


wb = openpyxl.load_workbook(excel_file_parse())
sheet = wb[sheet_in_excel()]
excel_req_info = []  # [ (row_nr, req), (row_nr, req), .. ]

# here you iterate over the rows in the specific column
for row in range(9, sheet.max_row + 1):
    for column in "A":  # Here you can add or reduce the columns
        cell_name = "{}{}".format(column, row)
        excel_req_info.append(
            (row, sheet[cell_name].value)
        )

tree = ET.ElementTree(file=xml_file_parse())
# tree = ET.ElementTree(file="C:/Users/User/Desktop/Temp/the_latest/Eng8_SW_Tests.xml")
root = tree.getroot()

idents = [el.text for el in root.iterfind('.//testcase/ident')]
descs = [el.text for el in root.iterfind('.//testcase/description')]

# pprint(idents)
# pprint(descs)

d = {id_: desc for id_, desc in zip(idents, descs)}
# print(d)
from typing import List, Dict, Union, Any


def find_trace(trace: str) -> List[str]:
    l = []
    for k, v in d.items():
        if trace in v:
            l.append(k)
    return l


for row_nr, req in excel_req_info:
    trace_id = find_trace(req)
    trace_id = str(trace_id)[2:-2]
    print(row_nr, req, trace_id)
    sheet.cell(row=row_nr, column=13).value = trace_id

wb.save('C:/Project/ReqParser/output/new_big_file.xlsx')

ConfType = Dict[str, Union[str, List[str]]]


















def get_files(conf_: ConfType) -> List[str]:
    return conf_['file_names']


def process(conf_: ConfType) -> Any:
    pass


def process(conf_: ConfType) -> Any:
    pass


def process(conf_: ConfType) -> Any:
    pass


def process(conf_: ConfType) -> Any:
    pass


def process(conf_: ConfType) -> Any:
    pass


def process(conf_: ConfType) -> Any:
    pass


def process(conf_: ConfType) -> Any:
    pass


if __name__ == '__main__':
    # python script.py conf.json
    # sys.argv = ['script.py', 'conf.json']
    json_parameter = sys.argv[1]

    conf = json.load(json_parameter)

    file_content = get_files(conf)
