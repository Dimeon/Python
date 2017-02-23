import openpyxl
import xml.etree.ElementTree as ET
import json
from typing import List


def get_excel_file_path() -> str:
    excel_filename = 'config.json'
    with open(excel_filename) as ex_f:
        pop_data = json.load(ex_f)
        for pop_dict in pop_data:
            path_name = pop_dict['Excel File']
            print(path_name)
            return path_name


def get_excel_sheet_name() -> str:
    excel_filename = 'config.json'
    with open(excel_filename) as ex_f:
        pop_data = json.load(ex_f)
        for pop_dict in pop_data:
            sheet_name = pop_dict['Excel Sheet']
            print(sheet_name)
            return sheet_name


def paths_to_xml_file() -> List[str]:
    xml_filename = 'xml_config.json'
    xml_path_str = []
    with open(xml_filename) as ex_f:
        pop_p_data = json.load(ex_f)
        for key, value in pop_p_data.items():
            xml_path_str.append(value)
    return xml_path_str


def parse_xml() -> str:
    ttt = paths_to_xml_file()
    idents = []
    descs = []
    for xml_file in ttt:
        tree = ET.ElementTree(file=xml_file)
        root = tree.getroot()

        idents += [el.text for el in root.iterfind('.//testcase/ident')]
        descs += [el.text for el in root.iterfind('.//testcase/description')]

    d = {id_: req for id_, req in zip(idents, descs)}

    return d


def find_trace(trace: str) -> List[str]:
    l = []
    d = parse_xml()
    for k, v in d.items():
        if trace in v:
            l.append(k)
    return l


wb = openpyxl.load_workbook(get_excel_file_path())
sheet = wb[get_excel_sheet_name()]
excel_req_info = []  # [ (row_nr, req), (row_nr, req), .. ]

# here you iterate over the rows in the specific column
for row in range(9, sheet.max_row + 1):
    for column in "A":  # Here you can add or reduce the columns
        cell_name = "{}{}".format(column, row)
        excel_req_info.append(
            (row, sheet[cell_name].value)
        )

for row_nr, req in excel_req_info:
    trace_id = find_trace(req)
    trace_id = str(trace_id)[2:-2]
    print(row_nr, req, trace_id)
    sheet.cell(row=row_nr, column=13).value = trace_id

wb.save('C:/Project/Python/ReqParser/output/new_big_file.xlsx')
